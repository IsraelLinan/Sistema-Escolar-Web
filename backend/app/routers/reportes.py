from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from datetime import date
import io
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.database import get_conn, put_conn
from app.auth import verificar_token

router = APIRouter(prefix="/reportes", tags=["Reportes"])


_QUERY_ASISTENCIA = """
    SELECT e.nombre, ie.hora_ingreso, ie.hora_salida, 'Estudiante' as tipo
    FROM ingresos_estudiantes ie
    JOIN estudiantes e ON e.id = ie.estudiante_id
    WHERE DATE(ie.hora_ingreso) = %s

    UNION ALL

    SELECT d.nombre, id2.hora_ingreso, id2.hora_salida, 'Docente' as tipo
    FROM ingresos_docentes id2
    JOIN docentes d ON d.id = id2.docente_id
    WHERE DATE(id2.hora_ingreso) = %s

    UNION ALL

    SELECT ac.nombre, ia.hora_ingreso, ia.hora_salida, 'Auxiliar' as tipo
    FROM ingresos_auxiliares ia
    JOIN auxiliares_codigos ac ON ac.id = ia.auxiliar_id
    WHERE DATE(ia.hora_ingreso) = %s

    ORDER BY hora_ingreso
"""


def _obtener_registros(cur, fecha_filtro):
    cur.execute(_QUERY_ASISTENCIA, (fecha_filtro, fecha_filtro, fecha_filtro))
    rows = cur.fetchall()
    return [
        {
            "nombre": r[0],
            "hora_ingreso": r[1].strftime('%H:%M:%S') if r[1] else None,
            "hora_salida": r[2].strftime('%H:%M:%S') if r[2] else None,
            "tipo": r[3]
        }
        for r in rows
    ]


@router.get("/asistencia")
def reporte_asistencia(fecha: str = None, tipo: str = None, pagina: int = 1, por_pagina: int = 20, usuario: str = Depends(verificar_token)):
    conn = get_conn()
    try:
        cur = conn.cursor()
        fecha_filtro = fecha if fecha else date.today().isoformat()

        registros = _obtener_registros(cur, fecha_filtro)
        cur.close()

        if tipo and tipo != 'Todos':
            registros = [r for r in registros if r['tipo'] == tipo]

        total = len(registros)
        inicio = (pagina - 1) * por_pagina
        fin = inicio + por_pagina
        registros_paginados = registros[inicio:fin]

        return {
            "fecha": fecha_filtro,
            "registros": registros_paginados,
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": (total + por_pagina - 1) // por_pagina
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        put_conn(conn)


@router.get("/exportar-excel")
def exportar_excel(fecha: str = None, usuario: str = Depends(verificar_token)):
    conn = get_conn()
    try:
        cur = conn.cursor()
        fecha_filtro = fecha if fecha else date.today().isoformat()
        registros = _obtener_registros(cur, fecha_filtro)
        cur.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Asistencia"

        ws.merge_cells('A1:D1')
        ws['A1'] = f'REPORTE DE ASISTENCIA - {fecha_filtro}'
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1E3A6E', end_color='1E3A6E', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        headers = ['Nombre', 'Tipo', 'Hora Ingreso', 'Hora Salida']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4f8ef7', end_color='4f8ef7', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        for row_idx, reg in enumerate(registros, 3):
            ws.cell(row=row_idx, column=1, value=reg['nombre'])
            ws.cell(row=row_idx, column=2, value=reg['tipo'])
            ws.cell(row=row_idx, column=3, value=reg['hora_ingreso'] or '—')
            ws.cell(row=row_idx, column=4, value=reg['hora_salida'] or '—')

            fill_color = 'F0F4FF' if row_idx % 2 == 0 else 'FFFFFF'
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type='solid'
                )
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=asistencia_{fecha_filtro}.xlsx'}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        put_conn(conn)


@router.get("/exportar-pdf")
def exportar_pdf(fecha: str = None, usuario: str = Depends(verificar_token)):
    conn = get_conn()
    try:
        cur = conn.cursor()
        fecha_filtro = fecha if fecha else date.today().isoformat()
        registros = _obtener_registros(cur, fecha_filtro)
        cur.close()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        title = Paragraph(f'<b>REPORTE DE ASISTENCIA — {fecha_filtro}</b>', styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))

        data = [['Nombre', 'Tipo', 'Hora Ingreso', 'Hora Salida']]
        for reg in registros:
            data.append([reg['nombre'], reg['tipo'], reg['hora_ingreso'] or '—', reg['hora_salida'] or '—'])

        table = Table(data, colWidths=[250, 100, 120, 120])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A6E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F0F4FF'), colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#2e3350')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type='application/pdf',
            headers={'Content-Disposition': f'attachment; filename=asistencia_{fecha_filtro}.pdf'}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        put_conn(conn)